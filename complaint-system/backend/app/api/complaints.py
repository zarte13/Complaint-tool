from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Response
from sqlalchemy.orm import Session
from sqlalchemy import desc, asc, or_, and_
from typing import List, Optional, Union
from datetime import datetime
import io
import csv
from app.database.database import get_db
from app.models.models import Complaint, Company, Part, ComplaintAttachment
from app.schemas.schemas import (
    ComplaintCreate, ComplaintResponse, ComplaintUpdate,
    AttachmentResponse, AttachmentUploadResponse,
    ComplaintSearchResponse
)
from app.utils.file_handler import save_upload_file, validate_file, delete_file
import mimetypes
import os
from app.auth.dependencies import require_admin

router = APIRouter(prefix="/api/complaints", tags=["complaints"])
# Register alias routes to support both "/api/complaints" and "/api/complaints/" without 307 redirects

# Accept both "" and "/" for collection POST
@router.post("", response_model=ComplaintResponse)
@router.post("/", response_model=ComplaintResponse)
async def create_complaint(
    complaint: ComplaintCreate,
    db: Session = Depends(get_db)
):
    """Create a new complaint"""
    # Verify company exists
    company = db.query(Company).filter(Company.id == complaint.company_id).first()
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    
    # Verify part exists
    part = db.query(Part).filter(Part.id == complaint.part_id).first()
    if not part:
        raise HTTPException(status_code=404, detail="Part not found")
    
    db_complaint = Complaint(**complaint.dict())
    db.add(db_complaint)
    db.commit()
    db.refresh(db_complaint)
    return db_complaint

# Accept both "" and "/" for collection GET
@router.get("", response_model=ComplaintSearchResponse)
@router.get("/", response_model=ComplaintSearchResponse)
async def get_complaints(
    page: int = Query(1, ge=1),
    size: int = Query(10, ge=1, le=100),
    search: Optional[str] = Query(None),
    status: Union[str, List[str], None] = Query(None, description="Single status or multiple statuses (comma-separated or repeated)"),
    issue_type: Optional[str] = Query(None),
    company_id: Optional[int] = Query(None),
    part_number: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    sort_by: Optional[str] = Query(None, regex=r"^(created_at|updated_at|company|part|status)$"),
    sort_order: Optional[str] = Query("desc", regex="^(asc|desc)$"),
    db: Session = Depends(get_db)
):
    """Get complaints with advanced filtering, search, and pagination"""
    query = db.query(Complaint).join(Company).join(Part).filter(Complaint.is_deleted == False)
    
    # Global search across multiple fields
    if search:
        search_term = f"%{search}%"
        # Handle ID search by converting to string for LIKE comparison
        try:
            # Try to convert search to int for exact ID match
            search_id = int(search)
            id_filter = Complaint.id == search_id
        except ValueError:
            # If search can't be converted to int, use false condition
            id_filter = False
        
        query = query.filter(
            or_(
                id_filter,
                Part.part_number.ilike(search_term),
                Part.description.ilike(search_term),
                Company.name.ilike(search_term),
                Complaint.details.ilike(search_term),
                Complaint.work_order_number.ilike(search_term),
                Complaint.occurrence.ilike(search_term),
                Complaint.part_received.ilike(search_term)
            )
        )
    
    # Status filter - support both single status and multiple statuses (comma-separated or repeated)
    if status:
        def _normalize_status_value(s: str) -> Optional[str]:
            s_lower = s.strip().lower()
            if not s_lower:
                return None
            if s_lower == "closed":
                return "resolved"
            if s_lower in {"open", "in_progress", "resolved"}:
                return s_lower
            return None

        if isinstance(status, str):
            raw_list = [seg for seg in (status.split(",") if "," in status else [status])]
        elif isinstance(status, list):
            raw_list = status
        else:
            raw_list = []

        norm_list: List[str] = []
        for seg in raw_list:
            norm = _normalize_status_value(str(seg))
            if norm:
                norm_list.append(norm)

        # Deduplicate while preserving order
        norm_list = list(dict.fromkeys(norm_list))

        if len(norm_list) == 1:
            query = query.filter(Complaint.status == norm_list[0])
        elif len(norm_list) > 1:
            query = query.filter(Complaint.status.in_(norm_list))
        # If no valid statuses after normalization, leave query unchanged
    
    # Issue type filter
    if issue_type:
        query = query.filter(Complaint.issue_type == issue_type)
    
    # Company filter
    if company_id:
        query = query.filter(Complaint.company_id == company_id)
    
    # Part number filter
    if part_number:
        query = query.filter(Part.part_number.ilike(f"%{part_number}%"))
    
    # Date range filter
    if date_from:
        try:
            date_from_dt = datetime.fromisoformat(date_from.replace('Z', '+00:00'))
            query = query.filter(Complaint.created_at >= date_from_dt)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date_from format")
    
    if date_to:
        try:
            date_to_dt = datetime.fromisoformat(date_to.replace('Z', '+00:00'))
            query = query.filter(Complaint.created_at <= date_to_dt)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date_to format")
    
    # Sorting
    sort_column_map = {
        'created_at': Complaint.created_at,
        'updated_at': Complaint.updated_at,
        'company': Company.name,
        'part': Part.part_number,
        'status': Complaint.status
    }
    
    if sort_by and sort_by in sort_column_map:
        sort_column = sort_column_map[sort_by]
        if sort_order == 'asc':
            query = query.order_by(asc(sort_column))
        else:
            query = query.order_by(desc(sort_column))
    else:
        query = query.order_by(desc(Complaint.created_at))
    
    # Pagination
    total = query.count()
    total_pages = (total + size - 1) // size
    
    complaints = query.offset((page - 1) * size).limit(size).all()
    
    return {
        "items": complaints,
        "pagination": {
            "page": page,
            "size": size,
            "total": total,
            "total_pages": total_pages
        }
    }

@router.get("/{complaint_id}", response_model=ComplaintResponse)
async def get_complaint(
    complaint_id: int,
    db: Session = Depends(get_db)
):
    """Get a specific complaint"""
    complaint = db.query(Complaint).filter(Complaint.id == complaint_id, Complaint.is_deleted == False).first()
    if not complaint:
        raise HTTPException(status_code=404, detail="Complaint not found")
    return complaint

@router.put("/{complaint_id}", response_model=ComplaintResponse)
async def update_complaint(
    complaint_id: int,
    complaint_update: ComplaintUpdate,
    db: Session = Depends(get_db)
):
    """Update a complaint"""
    complaint = db.query(Complaint).filter(Complaint.id == complaint_id, Complaint.is_deleted == False).first()
    if not complaint:
        raise HTTPException(status_code=404, detail="Complaint not found")
    
    update_data = complaint_update.dict(exclude_unset=True)
    # ComplaintUpdate.status is normalized in schema to canonical values ("open", "in_progress", "resolved")
    # Still defensively handle legacy "closed" if it appears
    if "status" in update_data and isinstance(update_data["status"], str):
        status_val = update_data["status"].strip().lower()
        if status_val == "closed":
            update_data["status"] = "resolved"
    for field, value in update_data.items():
        setattr(complaint, field, value)

    # Ensure default status is "open" if not explicitly set on creation elsewhere
    if not complaint.status:
        complaint.status = "open"
    
    db.commit()
    db.refresh(complaint)
    return complaint

@router.post("/{complaint_id}/attachments", response_model=AttachmentUploadResponse)
@router.post("/{complaint_id}/attachments/", response_model=AttachmentUploadResponse)
async def upload_attachment(
    complaint_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """Upload file attachment to a complaint"""
    # Verify complaint exists
    complaint = db.query(Complaint).filter(Complaint.id == complaint_id, Complaint.is_deleted == False).first()
    if not complaint:
        raise HTTPException(status_code=404, detail="Complaint not found")
    
    # Validate file
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")
    
    # Check file size
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:  # 10MB limit
        raise HTTPException(status_code=413, detail="File too large")
    
    # Reset file pointer
    await file.seek(0)
    
    # Get MIME type
    mime_type = file.content_type or mimetypes.guess_type(file.filename)[0] or "application/octet-stream"
    
    # Save file
    unique_filename, file_path = await save_upload_file(file, complaint_id)
    
    # Create attachment record
    attachment = ComplaintAttachment(
        complaint_id=complaint_id,
        filename=unique_filename,
        original_filename=file.filename,
        file_path=file_path,
        file_size=len(content),
        mime_type=mime_type
    )
    
    db.add(attachment)
    
    # Update complaint has_attachments flag
    complaint.has_attachments = True
    
    db.commit()
    db.refresh(attachment)
    
    return attachment

@router.get("/{complaint_id}/attachments", response_model=List[AttachmentResponse])
@router.get("/{complaint_id}/attachments/", response_model=List[AttachmentResponse])
async def get_attachments(
    complaint_id: int,
    db: Session = Depends(get_db)
):
    """Get all attachments for a complaint"""
    complaint = db.query(Complaint).filter(Complaint.id == complaint_id).first()
    if not complaint:
        raise HTTPException(status_code=404, detail="Complaint not found")
    
    attachments = db.query(ComplaintAttachment).filter(
        ComplaintAttachment.complaint_id == complaint_id
    ).all()
    
    return attachments

@router.get("/attachments/{attachment_id}/download")
@router.get("/attachments/{attachment_id}/download/")
async def download_attachment(
    attachment_id: int,
    db: Session = Depends(get_db)
):
    """Download a specific attachment file"""
    attachment = db.query(ComplaintAttachment).filter(
        ComplaintAttachment.id == attachment_id
    ).first()
    
    if not attachment:
        raise HTTPException(status_code=404, detail="Attachment not found")
    
    # Check if file exists
    if not os.path.exists(attachment.file_path):
        raise HTTPException(status_code=404, detail="File not found on server")
    
    # Get MIME type
    mime_type = attachment.mime_type or mimetypes.guess_type(attachment.original_filename)[0] or "application/octet-stream"
    
    # Read file content
    with open(attachment.file_path, "rb") as f:
        file_content = f.read()
    
    # Return file as response
    return Response(
        content=file_content,
        media_type=mime_type,
        headers={
            "Content-Disposition": f"attachment; filename=\"{attachment.original_filename}\""
        }
    )

@router.delete("/attachments/{attachment_id}")
async def delete_attachment(
    attachment_id: int,
    db: Session = Depends(get_db)
):
    """Delete an attachment"""
    attachment = db.query(ComplaintAttachment).filter(
        ComplaintAttachment.id == attachment_id
    ).first()
    
    if not attachment:
        raise HTTPException(status_code=404, detail="Attachment not found")
    
    # Delete file from filesystem
    delete_file(attachment.file_path)
    
    # Delete from database
    db.delete(attachment)
    
    # Check if complaint has any remaining attachments
    remaining = db.query(ComplaintAttachment).filter(
        ComplaintAttachment.complaint_id == attachment.complaint_id
    ).count()
    
    if remaining == 0:
        # Update complaint has_attachments flag
        complaint = db.query(Complaint).filter(
            Complaint.id == attachment.complaint_id
        ).first()
        if complaint:
            complaint.has_attachments = False
    
    db.commit()
    return {"message": "Attachment deleted successfully"}

@router.get("/export/csv")
async def export_csv(
    search: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    issue_type: Optional[str] = Query(None),
    company_id: Optional[int] = Query(None),
    part_number: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    """Export complaints to CSV format"""
    query = db.query(Complaint).join(Company).join(Part)
    
    # Apply same filters as get_complaints
    if search:
        search_term = f"%{search}%"
        # Handle ID search by converting to string for LIKE comparison
        try:
            # Try to convert search to int for exact ID match
            search_id = int(search)
            id_filter = Complaint.id == search_id
        except ValueError:
            # If search can't be converted to int, use false condition
            id_filter = False
        
        query = query.filter(
            or_(
                id_filter,
                Part.part_number.ilike(search_term),
                Part.description.ilike(search_term),
                Company.name.ilike(search_term),
                Complaint.details.ilike(search_term),
                Complaint.work_order_number.ilike(search_term),
                Complaint.occurrence.ilike(search_term),
                Complaint.part_received.ilike(search_term)
            )
        )
    
    if status:
        # Normalize "closed" synonym to canonical "resolved"
        norm_status = status.lower()
        if norm_status == "closed":
            norm_status = "resolved"
        query = query.filter(Complaint.status == norm_status)
    
    if issue_type:
        query = query.filter(Complaint.issue_type == issue_type)
    
    if company_id:
        query = query.filter(Complaint.company_id == company_id)
    
    if part_number:
        query = query.filter(Part.part_number.ilike(f"%{part_number}%"))
    
    if date_from:
        try:
            date_from_dt = datetime.fromisoformat(date_from.replace('Z', '+00:00'))
            query = query.filter(Complaint.created_at >= date_from_dt)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date_from format")
    
    if date_to:
        try:
            date_to_dt = datetime.fromisoformat(date_to.replace('Z', '+00:00'))
            query = query.filter(Complaint.created_at <= date_to_dt)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date_to format")
    
    complaints = query.order_by(desc(Complaint.created_at)).all()
    
    # Create CSV
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['ID', 'Company', 'Part Number', 'Issue Type', 'Status', 'Created At', 'Details', 'Work Order', 'Occurrence', 'Part Received'])
    
    for complaint in complaints:
        writer.writerow([
            complaint.id,
            complaint.company.name,
            complaint.part.part_number,
            complaint.issue_type,
            complaint.status,
            complaint.created_at.isoformat(),
            complaint.details,
            complaint.work_order_number,
            complaint.occurrence or '',
            complaint.part_received or ''
        ])
    
    return Response(
        content=output.getvalue(),
        media_type='text/csv',
        headers={'Content-Disposition': 'attachment; filename=complaints.csv'}
    )

@router.get("/export/excel")
async def export_excel(
    search: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    issue_type: Optional[str] = Query(None),
    company_id: Optional[int] = Query(None),
    part_number: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    """Export complaints to Excel format"""
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="Excel export not available. Please install openpyxl: pip install openpyxl")
    
    query = db.query(Complaint).join(Company).join(Part)
    
    # Apply same filters as get_complaints
    if search:
        search_term = f"%{search}%"
        # Handle ID search by converting to string for LIKE comparison
        try:
            # Try to convert search to int for exact ID match
            search_id = int(search)
            id_filter = Complaint.id == search_id
        except ValueError:
            # If search can't be converted to int, use false condition
            id_filter = False
        
        query = query.filter(
            or_(
                id_filter,
                Part.part_number.ilike(search_term),
                Part.description.ilike(search_term),
                Company.name.ilike(search_term),
                Complaint.details.ilike(search_term),
                Complaint.work_order_number.ilike(search_term),
                Complaint.occurrence.ilike(search_term),
                Complaint.part_received.ilike(search_term)
            )
        )
    
    if status:
        # Normalize "closed" synonym to canonical "resolved"
        norm_status = status.lower()
        if norm_status == "closed":
            norm_status = "resolved"
        query = query.filter(Complaint.status == norm_status)
    
    if issue_type:
        query = query.filter(Complaint.issue_type == issue_type)
    
    if company_id:
        query = query.filter(Complaint.company_id == company_id)
    
    if part_number:
        query = query.filter(Part.part_number.ilike(f"%{part_number}%"))
    
    if date_from:
        try:
            date_from_dt = datetime.fromisoformat(date_from.replace('Z', '+00:00'))
            query = query.filter(Complaint.created_at >= date_from_dt)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date_from format")
    
    if date_to:
        try:
            date_to_dt = datetime.fromisoformat(date_to.replace('Z', '+00:00'))
            query = query.filter(Complaint.created_at <= date_to_dt)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date_to format")
    
    complaints = query.order_by(desc(Complaint.created_at)).all()
    
    # Create Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Complaints"
    
    # Headers
    headers = ['ID', 'Company', 'Part Number', 'Issue Type', 'Status', 'Created At', 'Details', 'Work Order', 'Occurrence', 'Part Received']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Data
    for row, complaint in enumerate(complaints, 2):
        ws.cell(row=row, column=1, value=complaint.id)
        ws.cell(row=row, column=2, value=complaint.company.name)
        ws.cell(row=row, column=3, value=complaint.part.part_number)
        ws.cell(row=row, column=4, value=complaint.issue_type)
        ws.cell(row=row, column=5, value=complaint.status)
        ws.cell(row=row, column=6, value=complaint.created_at.isoformat())
        ws.cell(row=row, column=7, value=complaint.details)
        ws.cell(row=row, column=8, value=complaint.work_order_number)
        ws.cell(row=row, column=9, value=complaint.occurrence or '')
        ws.cell(row=row, column=10, value=complaint.part_received or '')
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return Response(
        content=buffer.getvalue(),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=complaints.xlsx'}
    )

@router.delete("/{complaint_id}", status_code=204)
@router.delete("/{complaint_id}/", status_code=204)
async def delete_complaint(
    complaint_id: int,
    db: Session = Depends(get_db),
    _admin = Depends(require_admin),
):
    """Admin-only: Soft delete a complaint (hide from UI), keep data and attachments."""
    complaint = db.query(Complaint).filter(Complaint.id == complaint_id, Complaint.is_deleted == False).first()
    if not complaint:
        raise HTTPException(status_code=404, detail="Complaint not found")

    complaint.is_deleted = True
    db.add(complaint)
    db.commit()
    return Response(status_code=204)